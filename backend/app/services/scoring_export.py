"""Build Excel exports for entry rosters vs points and per-golfer point ledger."""
from __future__ import annotations

from collections import defaultdict
from io import BytesIO
from typing import Any, DefaultDict, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.models import BonusPoint, DailyScore, Entry, Player, Tournament


def _player_name_map(db: Session, player_ids: List[str]) -> Dict[str, str]:
    ids = [p for p in player_ids if p]
    if not ids:
        return {}
    rows = db.query(Player).filter(Player.player_id.in_(ids)).all()
    return {r.player_id: (r.full_name or f"{r.first_name} {r.last_name}".strip()) for r in rows}


def _slot_player_ids(entry: Entry) -> List[Optional[str]]:
    return [
        entry.player1_id,
        entry.player2_id,
        entry.player3_id,
        entry.player4_id,
        entry.player5_id,
        entry.player6_id,
    ]


def build_workbook(db: Session, tournament_id: int) -> Tuple[Workbook, str]:
    """
    Sheet 1: one row per pool entry — participant, each roster slot name + that golfer's
    total points for this entry (base + bonuses attributed to them), team bonus, entry total.
    Sheet 2: one row per point line — golfer, entry, round, kind, detail, points.
    """
    tournament = db.query(Tournament).filter(Tournament.id == tournament_id).first()
    if not tournament:
        raise ValueError("Tournament not found")

    entries = (
        db.query(Entry)
        .filter(Entry.tournament_id == tournament_id)
        .order_by(Entry.id.asc())
        .all()
    )
    entry_by_id: Dict[int, Entry] = {e.id: e for e in entries}

    official_entry_total: Dict[int, float] = {}
    if entries:
        q_tot = (
            db.query(DailyScore.entry_id, func.coalesce(func.sum(DailyScore.total_points), 0.0))
            .filter(DailyScore.entry_id.in_([e.id for e in entries]))
            .group_by(DailyScore.entry_id)
            .all()
        )
        official_entry_total = {int(eid): round(float(t), 4) for eid, t in q_tot}

    # --- Collect all player ids for name lookup
    all_pids: List[str] = []
    for e in entries:
        for pid in _slot_player_ids(e):
            if pid:
                all_pids.append(str(pid))
    bp_rows = (
        db.query(BonusPoint)
        .join(Entry, BonusPoint.entry_id == Entry.id)
        .filter(Entry.tournament_id == tournament_id)
        .all()
    )
    for bp in bp_rows:
        if bp.player_id:
            all_pids.append(str(bp.player_id))
    name_by_id = _player_name_map(db, list(set(all_pids)))

    # --- Per entry: base points by player_id from daily_scores.details
    entry_base_by_player: Dict[int, DefaultDict[str, float]] = {}
    for e in entries:
        entry_base_by_player[e.id] = defaultdict(float)

    daily_all = (
        db.query(DailyScore)
        .join(Entry, DailyScore.entry_id == Entry.id)
        .filter(Entry.tournament_id == tournament_id)
        .order_by(DailyScore.entry_id, DailyScore.round_id)
        .all()
    )

    for ds in daily_all:
        details = ds.details or {}
        bb = details.get("base_breakdown") or {}
        tgt = entry_base_by_player.setdefault(ds.entry_id, defaultdict(float))
        for i in range(1, 7):
            cell = bb.get(f"player{i}") or {}
            pid = cell.get("player_id")
            if not pid:
                continue
            tgt[str(pid)] += float(cell.get("points") or 0.0)

    # --- Bonus points by entry and player_id (including None = team / entry-level)
    entry_bonus_by_player: Dict[int, DefaultDict[str, float]] = defaultdict(lambda: defaultdict(float))
    for bp in bp_rows:
        key = str(bp.player_id) if bp.player_id else "__team__"
        entry_bonus_by_player[bp.entry_id][key] += float(bp.points or 0.0)

    def total_for_player_on_entry(entry_id: int, player_id: Optional[str]) -> float:
        if not player_id:
            return 0.0
        pid = str(player_id)
        b = entry_base_by_player.get(entry_id, defaultdict(float)).get(pid, 0.0)
        bon = entry_bonus_by_player.get(entry_id, defaultdict(float)).get(pid, 0.0)
        return round(b + bon, 4)

    def team_bonus_for_entry(entry_id: int) -> float:
        return round(entry_bonus_by_player.get(entry_id, defaultdict(float)).get("__team__", 0.0), 4)

    # --- Sheet 1
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Entries & golfer points"

    headers1 = [
        "Participant",
        "Email",
        "Entry ID",
        "Player 1 name",
        "Player 1 ID",
        "P1 points (base+bonus)",
        "Player 2 name",
        "Player 2 ID",
        "P2 points",
        "Player 3 name",
        "Player 3 ID",
        "P3 points",
        "Player 4 name",
        "Player 4 ID",
        "P4 points",
        "Player 5 name",
        "Player 5 ID",
        "P5 points",
        "Player 6 name",
        "Player 6 ID",
        "P6 points",
        "Team / entry bonus (no golfer)",
        "Entry total points",
    ]
    ws1.append(headers1)
    for c in range(1, len(headers1) + 1):
        ws1.cell(row=1, column=c).font = Font(bold=True)

    for e in entries:
        part = e.participant
        slots = _slot_player_ids(e)
        row_out: List[Any] = [
            part.name if part else "",
            part.email if part else "",
            e.id,
        ]
        for pid in slots:
            if not pid:
                row_out.extend(["", "", 0.0])
                continue
            nm = name_by_id.get(str(pid), "")
            pts = total_for_player_on_entry(e.id, pid)
            row_out.extend([nm, str(pid), pts])
        tb = team_bonus_for_entry(e.id)
        row_out.append(tb)
        row_out.append(official_entry_total.get(e.id, 0.0))
        ws1.append(row_out)

    # --- Sheet 2: ledger
    ws2 = wb.create_sheet("Golfer point ledger")

    h2 = [
        "Golfer name",
        "Golfer ID",
        "Pool participant",
        "Entry ID",
        "Round",
        "Source",
        "Detail",
        "Points",
    ]
    ws2.append(h2)
    for c in range(1, len(h2) + 1):
        ws2.cell(row=1, column=c).font = Font(bold=True)

    for ds in daily_all:
        entry = entry_by_id.get(ds.entry_id)
        if not entry:
            continue
        pname = entry.participant.name if entry.participant else ""
        details = ds.details or {}
        bb = details.get("base_breakdown") or {}
        for i in range(1, 7):
            cell = bb.get(f"player{i}") or {}
            pid = cell.get("player_id")
            if not pid:
                continue
            pts = float(cell.get("points") or 0.0)
            pos = cell.get("position")
            st = cell.get("status")
            detail = f"Position points (slot {i}); leaderboard pos={pos}, status={st}"
            ws2.append(
                [
                    name_by_id.get(str(pid), ""),
                    str(pid),
                    pname,
                    entry.id,
                    ds.round_id,
                    "base_position",
                    detail,
                    pts,
                ]
            )

    for bp in sorted(bp_rows, key=lambda x: (x.player_id or "", x.entry_id, x.round_id, x.id)):
        entry = entry_by_id.get(bp.entry_id)
        pname = entry.participant.name if entry and entry.participant else ""
        pid = bp.player_id
        gname = name_by_id.get(str(pid), "") if pid else ""
        hole = bp.hole
        detail = f"type={bp.bonus_type}"
        if hole is not None:
            detail += f", hole={hole}"
        ws2.append(
            [
                gname if pid else "(team / entry)",
                str(pid) if pid else "",
                pname,
                bp.entry_id,
                bp.round_id,
                f"bonus:{bp.bonus_type}",
                detail,
                float(bp.points or 0),
            ]
        )

    fn = f"eldorado_scoring_{tournament.name.replace(' ', '_')}_{tournament.year}_export.xlsx"
    safe = "".join(c for c in fn if c.isalnum() or c in "._-")
    return wb, safe


def workbook_to_bytes(wb: Workbook) -> BytesIO:
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
